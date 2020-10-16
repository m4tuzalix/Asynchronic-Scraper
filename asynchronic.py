from requests_html import AsyncHTMLSession
import re
from openpyxl import Workbook, load_workbook

class Scraper(AsyncHTMLSession):
    def __init__(self):
        AsyncHTMLSession.__init__(self)
        self.comments = []
        self.scraped_pages = []
    
    async def __comments(self, links):
        for link in links:
            try:
                link_response = await self.get(link, verify=False)
                link_comments = link_response.html.find("div[class='text'] p")
                for comment in link_comments:
                    try:
                        r = re.search(r'(?:: )(.*)', comment.text)
                        self.comments.append(r.group(0)[1:])
                    except AttributeError:
                        self.comments.append(comment.text)
            except:
                continue

    async def _base(self, range1, range2):
        for x in range(range1, range2):
            try:
                site_content = await self.get(f"https://www.wykop.pl/strona/{x}/", verify=False)
                topics = site_content.html.find("div[class='lcontrast m-reset-margin'] h2 a")
                links = ([topic.attrs["href"] for topic in topics])
                await self.__comments(links)
            except:
                continue

class Excel(Workbook):
    def __init__(self):
        Workbook.__init__(self)
        self.ws = self.active
        self.workbook_name = f"list.xlsx"
        try:
            self.load = load_workbook(self.workbook_name)
        except FileNotFoundError:
            headers = "comment"
            self.ws.cell(row=1, column=1, value=headers) #// +1 because it must be at least 1
            self.save(self.workbook_name)
      
    def add_data(self, comments):
        self.load = load_workbook(self.workbook_name)
        self.ns = self.load.get_sheet_by_name("Sheet")
        for comment in comments:
            try:
                if len(comment) > 0:
                    data_row = [comment]
                    self.ns.append(data_row)
            except:
                continue
        self.load.save(self.workbook_name)

class Tasks(Scraper, Excel):
    def __init__(self, iterations=5):
        self.iterations = iterations
        Scraper.__init__(self)
        Excel.__init__(self)
    
    def task_creator(self):
        """
        Creates dynamic async functions and store them in 
        array whichc is unziped to asyncio loop starter
        """
        functions = []
        first = 3000
        second = 3300
        for x in range(self.iterations):
            async def base(range1=first,range2=second):
                await self._base(range1, range2)
            functions.append(base)
            if first and second < 6000:
                first += 300
                second += 300
            else: break
        return functions

    def start(self):
        self.run(*self.task_creator())
        self.add_data(self.comments)
    
if __name__ == "__main__":
    T = Tasks(iterations=8)
    T.start()
